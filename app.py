#app.py

from fastapi import FastAPI, HTTPException, UploadFile
from fastapi.responses import FileResponse
from pydantic import BaseModel
from urllib.parse import quote
from pymongo.mongo_client import MongoClient
from dotenv import load_dotenv
import requests
import xml.etree.ElementTree as ET
import json
import openpyxl
import datetime
import io
import time 
import os
import uvicorn
import webbrowser
import re
from openai import OpenAI
import pandas as pd
load_dotenv()
app = FastAPI()

API_KEY = os.getenv("API_KEY")
MONGO_USER = os.getenv("MONGO_USER")
MONGO_PASSWORD = os.getenv("MONGO_PASSWORD")

class AddressInput(BaseModel):
    address: str

uri = f"mongodb+srv://{MONGO_USER}:{MONGO_PASSWORD}@cluster0.j1ypizq.mongodb.net/?retryWrites=true&w=majority"
    # Create a new client and connect to the server
client = MongoClient(uri)
    # Send a ping to confirm a successful connection
db = client.get_database("AddressCheck")
collection = db["MyCollection"]


@app.get('/')
def hello_world():
    return "Hello,World"


@app.post("/uploadfile/")
async def create_upload_file(file: UploadFile):
    start = time.time()
    # check file 
    if file.filename.endswith('.xlsx'):
        f = await file.read()
        xlsx = io.BytesIO(f)
        wb = openpyxl.load_workbook(xlsx)
        ws = wb.active
        col_name_list = ['序號','地址','編號','站所四碼代號','站所名稱','站所簡碼','郵遞區號','ＳＤ註區','ＳＤ疊區','ＤＤ註區','ＤＤ疊區','ＭＤ註區','ＭＤ疊區','低溫註區','低溫疊區','聯運費用(起碼)','聯運費用(百公斤)','PUTAO_FLAG','新集貨註區','PUTCODE1','PUTCODE2','PUTCODE3','PUTCODE4','PUTCODE5','PUTCODE6','類別','到著簡碼','衛星區','註區','疊區','QRcode','聯運區提醒(*表示有聯運區,空白表示沒有)','MD 到著碼衛星區','MD 到著碼註區','MD 到著碼疊區','訊息']
        address_list = []
        for index, value in enumerate(col_name_list, 1):
            ws.cell(row=1, column=index, value=value)
        print(ws.max_row)
        for cells in ws['B2':'B' + str(ws.max_row)]:
            for cell in cells:
                if cell.value is not None:
                    cleaned_address = clean_address(cell.value)
                    address_list.append(cleaned_address)
                else:
                    address_list.append("")

        # call hsinchu api
        soap_response = send_soap_request_by_address_list(address_list)
        # write to excel
        response = parse_soap_response_by_list_uploadfile(wb, soap_response)
        print(response)
        if response != True:
            raise HTTPException(404, response)
        current_time = datetime.datetime.now()
        
        filename =  "地址比對" + str(current_time.year) + '-' + str(current_time.month) + '-' + str(current_time.day) + '-' + str(current_time.hour) + '-' + str(current_time.minute) + '-' + str(current_time.second) + ".xlsx"
        print(filename)
        wb.save('new.xlsx')
    else:
        raise EOFError
    # return None
    print(time.time()-start)
    return FileResponse('new.xlsx', headers={"Content-Disposition": f"attachment; filename={quote(filename)}"})

@app.post("/address_translate/")
async def address_translate(file: UploadFile):
    start = time.time()
    zipcode_dict = post_code()
    
    
    # check file 
    if file.filename.endswith('.xlsx'):
        f = await file.read()
        xlsx = io.BytesIO(f)
        wb = openpyxl.load_workbook(xlsx)
        ws = wb.active
        col_name_list = ['分號','城市','地址','郵遞區號','對應縣市區域','中文地址','編號','站所四碼代號','站所名稱','站所簡碼','郵遞區號','ＳＤ註區','ＳＤ疊區','ＤＤ註區','ＤＤ疊區','ＭＤ註區','ＭＤ疊區','低溫註區','低溫疊區','聯運費用(起碼)','聯運費用(百公斤)','PUTAO_FLAG','新集貨註區','PUTCODE1','PUTCODE2','PUTCODE3','PUTCODE4','PUTCODE5','PUTCODE6','類別','到著簡碼','衛星區','註區','疊區','QRcode','聯運區提醒(*表示有聯運區,空白表示沒有)','MD 到著碼衛星區','MD 到著碼註區','MD 到著碼疊區','訊息']
        address_list = []
        zipcode_list = []
        for index, value in enumerate(col_name_list, 1):
            ws.cell(row=1, column=index, value=value)
        # print(ws.max_row)
        for cells_a, cells_b, cells_c, cells_d in zip(ws['A2':'A' + str(ws.max_row)], ws['B2':'B' + str(ws.max_row)], ws['C2':'C' + str(ws.max_row)], ws['D2':'D' + str(ws.max_row)]):
            for cell_a, cell_b, cell_c, cell_d in zip(cells_a, cells_b, cells_c, cells_d):
                if cell_a.value is None:
                    break
                if cell_d.value is not None:
                    str_cell_d_value = str(cell_d.value)
                    if len(str(str_cell_d_value)) >= 3:
                        if is_valid_int(str_cell_d_value):
                            if int(str_cell_d_value[0:3]) in zipcode_dict:
                                zipcode = zipcode_dict[int(str_cell_d_value[0:3])]
                            else:
                                zipcode = ""
                        else:
                            zipcode = ""
                    else:
                        zipcode = ""
                else:
                    zipcode = ""
                zipcode_list.append(zipcode)

                if cell_b.value is not None:
                    # 合併地址
                    if zipcode != "":
                        cleaned_address = clean_address(zipcode) + "@" + clean_address(cell_c.value)
                    else:    
                        cleaned_address = clean_address(re.sub(r'[0-9]+', '', cell_b.value)) + "@" + clean_address(cell_c.value)
                    address_list.append(cleaned_address)
                elif cell_c.value is not None:
                    cleaned_address = clean_address(cell_c.value)
                    address_list.append(cleaned_address)
                else:
                    address_list.append("無法翻譯")
        # print(address_list)
        print("地址翻譯")
        address_list = gpt_translate(address_list)
        print("修復地址")
        address_list = fix_address(address_list)
        print("新竹物流api")
        soap_response = send_soap_request_by_address_list(address_list)
        print("寫入檔案")
        response = parse_soap_response_by_list(wb, soap_response, address_list, zipcode_list)
        print(response)
        if response != True:
            raise HTTPException(404, response)
        current_time = datetime.datetime.now()
        
        filename =  "地址比對" + str(current_time.year) + '-' + str(current_time.month) + '-' + str(current_time.day) + '-' + str(current_time.hour) + '-' + str(current_time.minute) + '-' + str(current_time.second) + ".xlsx"
        print(filename)
        wb.save('new.xlsx')
    else:
        raise EOFError
    # return None
    print(time.time()-start)
    return FileResponse('new.xlsx', headers={"Content-Disposition": f"attachment; filename={quote(filename)}"})
    

def post_code():
    # 取得目前資料夾路徑
    current_directory = os.getcwd()

    # 確認目前資料夾下是否存在"郵遞區號對照檔.xlsx"
    file_name = "郵遞區號對照檔.xlsx"
    file_path = os.path.join(current_directory, file_name)

    zipcode_dict = {}  # 儲存郵遞區號與縣市區域的對應關係

    if os.path.exists(file_path):
        # 使用 pandas 讀取 Excel 檔案
        df = pd.read_excel(file_path)

        # 從第二列開始讀取資料，並將A行對應B行放入字典
        for index, row in df.iterrows():
            zipcode = row['郵遞區號']
            area = row['縣市區域']
            zipcode_dict[zipcode] = area
        return zipcode_dict
    else:
        return False
    
def clean_address(address):
    # Use regular expression to remove special characters
    cleaned_address = re.sub(r'[\\/.,~!+?#&=%\n"]', '', address)
    return cleaned_address

def send_soap_request_by_address_list(address_list):
    request_body = ''
    before_request = '{"USER": "0494227", "NO": "", "ADDR": "'
    after_request = '", "ESDATE": "", "TEL": "", "TEL2": "", "EPRDCL": "", "EMARK": "", "ESCSNO": "", "EJAMT": "", "EQAMT": "", "EQAMTTYPE": "", "ELAMTTYPE": ""},'
    for address in address_list:
        request_body = request_body + before_request + address + after_request
    url = "https://is1fax.hct.com.tw/Webedi_Erstno_NEW2/WS_addrCompare1.asmx"
    headers = {
        'content-type': 'text/xml; charset=utf-8',
        'SOAPAction': 'http://tempuri.org/addrCompare_Json',
        'Content-Length': '<calculated when request is sent>'
    }
 
    xml_request = f"""<?xml version="1.0" encoding="utf-8"?>
    <soap:Envelope xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema" xmlns:soap="http://schemas.xmlsoap.org/soap/envelope/">
        <soap:Body>
            <addrCompare_Json xmlns="http://tempuri.org/">
                <Json>[{request_body}]</Json>
            </addrCompare_Json>
        </soap:Body>
    </soap:Envelope>
    """
    encoded_xml_request = xml_request.encode('utf-8')

    response = requests.post(url, data=encoded_xml_request, headers=headers)
    return response.text

def parse_soap_response_by_list(wb, soap_response, address_list, zipcode_list):
    ws = wb.active
    root = ET.fromstring(soap_response)[0][0]
    if root[0].tag =='faultcode':
        return root[1].text
    root = ET.fromstring(soap_response)[0][0][0]#讀取addrCompare_JsonResult層
    json_text = root.text
    if json_text is None:
        return "新竹物流API回傳為空"
    data = json.loads(json_text)
    row = 2
    address_column = ord('F')-ord('A')+1
    for i, address, zipcode in zip(data,address_list, zipcode_list):
        ws.cell(row = row, column = address_column-1, value = zipcode)
        ws.cell(row = row, column = address_column, value = address)
        index = ord('G')-ord('A')+1
        for k, v in i.items():
            ws.cell(row = row, column = index, value = v)
            index = index +1
        row = row + 1
    return True

def parse_soap_response_by_list_uploadfile(wb, soap_response):
    ws = wb.active
    root = ET.fromstring(soap_response)[0][0]
    if root[0].tag =='faultcode':
        return root[1].text
    root = ET.fromstring(soap_response)[0][0][0]#讀取addrCompare_JsonResult層
    json_text = root.text
    if json_text is None:
        return "新竹物流API回傳為空"
    data = json.loads(json_text)
    row = 2
    for i in data:
        index = ord('C')-ord('A')+1
        for k, v in i.items():
            ws.cell(row = row, column = index, value = v)
            index = index +1
        row = row + 1
    return True

def gpt_translate(address_list):
    sub_lists = []  # 用於存儲分割後的子列表
    response_list = []
    total = 0

    # 遍歷原始列表，每20個元素分割成一個子列表
    for i in range(0, len(address_list), 5):
        sub_lists.append(address_list[i:i+5])
    for sub in sub_lists:
        sysMsg = '''
        You are a Taiwan shipping address translator and the user will provide an list of addresses separated by "###". Each element consists of two parts. The first part is the city name, and the second part is the detailed address. The two parts are separated by '@'.
        Address example: "Kaohsiung City@No. 2,Lane 19, Wenya Street 5F###Zhubei City@7F, No. 66-3, Zhuangjing 3rd Rd.###TAINAN CITY@No. 387 Sec. 2, Yonghua Road 22/F A5###"

        The standard format of Taiwan postal address is:
        縣市+區+鄉鎮+街/路+號+樓+室的中文地址，且RM=ROOM=室 
        依照順序組合，並且"不包含重複元素"。
        Please translate each address into "Traditional Chinese" in sequence according to Taiwan's standard address format.

        You need to perform the following three steps:
        Step 1: Cut the address list provided by the user into all addresses according to "###".
        Step 2: Translate each address into "Traditional Chinese" according to the "Taiwan Postal Address Standard Format",如果遇到數字，請保留為數字，不要更改為中文數字，翻譯完成後，請對地址做地址正規化，,如果地址為空或是無法翻譯，則翻譯成"無法翻譯",The address must be translated regardless of whether it has been translated before.
        Step 3: The converted addresses are separated by ###. Each address cannot contain characters other than Chinese characters, numbers and "-".
        The rules you must follow are: Your answer should be a string with the following format: "address1###address2###address3###address4###”, without any explanation.
        '''

        prompt = '''
        address list data:
        {data}
        '''
        content = ''
        for index, address in enumerate(sub):
            content += address +'###'

        client = OpenAI(api_key=API_KEY)
        sub_len = len(sub)
        # print("sub_len: ", sub_len)
        gpt_len = 0
        while(gpt_len != sub_len):
            response = client.chat.completions.create(
            model="gpt-3.5-turbo-0125",
            messages=[
                {"role": "system", "content": sysMsg},
                {"role": "user", "content": content}
            ]
            )
            filtered_list = [x.replace('@', '').replace('\n', '').replace("TW", '').replace(' ', '').replace('[機械翻譯]','') for x in response.choices[0].message.content.split("###") if x]
            gpt_len = len(filtered_list)
            
            # print("gpt_len: ", gpt_len)
        total = total + gpt_len
        print("已翻譯",total,'/',str(len(address_list)))
        # print(filtered_list)
        response_list.extend(filtered_list)
    return response_list

def fix_address(address_list):
    # 取得目前資料夾路徑
    current_directory = os.getcwd()

    # 確認目前資料夾下是否存在"郵遞區號對照檔.xlsx"
    file_name = "變更地址.xlsx"
    file_path = os.path.join(current_directory, file_name)

    fixWord_dict = {}  # 儲存郵遞區號與縣市區域的對應關係

    if os.path.exists(file_path):
        # 使用 pandas 讀取 Excel 檔案
        df = pd.read_excel(file_path)

        # 從第二列開始讀取資料，並將A行對應B行放入字典
        for index, row in df.iterrows():
            errorWord = row['原']
            correctWord = row['改']
            fixWord_dict[errorWord] = correctWord
        for i, string in enumerate(address_list):
            for key, value in fixWord_dict.items():
                if key in string:
                    print("key",key)
                    print("i=",i," address= ",address_list[i])
                    address_list[i] = string.replace(key, str(value))
        return address_list
    else:
        return address_list
    
# 检查字符串是否可以转换为整数
def is_valid_int(s):
    # 使用isdigit()方法检查字符串是否只包含数字字符
    return s.isdigit()

if __name__ == '__main__':
    import uvicorn
    uvicorn.run(app)